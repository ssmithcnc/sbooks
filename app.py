from __future__ import annotations

import csv
import io
import base64
import html
import mimetypes
import os
import re
import smtplib
import ssl
import struct
import subprocess
import uuid
import zipfile
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path
from typing import List, Optional, Tuple
from urllib.parse import quote

from flask import Flask, jsonify, request, render_template, send_file, send_from_directory

from db import init_db, db

APP_TITLE = "S-Books"
FIRST_OF_MONTH_MAX_DAY = 15
SBOOKS_BRAND_ASSET = Path("static") / "sbooks-brand-badge.png"

def iso(d: date) -> str:
  return d.isoformat()

def parse_date(s: str) -> date:
  """Parse a date from ISO (YYYY-MM-DD) or common US format (MM/DD/YYYY)."""
  s = (s or "").strip()
  try:
    return date.fromisoformat(s)
  except Exception:
    try:
      return datetime.strptime(s, "%m/%d/%Y").date()
    except Exception:
      # Re-raise a clear error for API callers
      raise ValueError(f"Invalid date: {s!r}")

def now_iso() -> str:
  return datetime.now().replace(microsecond=0).isoformat()


BUSINESS_SETTINGS_DEFAULTS = {
  "company_name": "My Business",
  "company_address": "",
  "company_phone": "",
  "company_email": "",
  "company_logo_path": r"C:\Users\smith\Downloads\CNCPOWDER - Sbooks.jpg",
  "company_website": "https://cncpowder.com",
  "smtp_provider": "custom",
  "smtp_host": "",
  "smtp_port": "587",
  "smtp_username": "",
  "smtp_password": "",
  "smtp_use_tls": "1",
  "smtp_from_name": "",
  "invoice_payment_url_base": "",
  "invoice_prefix": "INV-",
  "estimate_prefix": "EST-",
  "next_invoice_number": "1001",
  "next_estimate_number": "1001",
  "default_tax_rate": "0",
  "default_terms": "Due on receipt",
}


def parse_float(value, default: float = 0.0) -> float:
  if value in (None, ""):
    return default
  if isinstance(value, (int, float)):
    return float(value)
  text = str(value).strip().replace("$", "").replace(",", "")
  if text.endswith("%"):
    text = text[:-1]
  try:
    return float(text)
  except Exception:
    return default


def parse_bool(value, default: bool = False) -> bool:
  if isinstance(value, bool):
    return value
  text = str(value or "").strip().lower()
  if text in {"1", "true", "yes", "y", "taxable", "active"}:
    return True
  if text in {"0", "false", "no", "n", "non-taxable", "inactive"}:
    return False
  return default


def normalize_header(value: str) -> str:
  return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def pick_value(row: dict, *keys, default=None):
  for key in keys:
    if key in row:
      val = row[key]
      if val not in (None, ""):
        return val
  return default


def clean_text(value, default: str = "") -> str:
  return str(value or default).strip()


def looks_like_report_footer(value: str) -> bool:
  text = clean_text(value).lower()
  if not text:
    return False
  if text == "this report contains no data.":
    return True
  if re.match(r"^(monday|tuesday|wednesday|thursday|friday|saturday|sunday),\s", text):
    return True
  return False


def json_error(message: str, status: int = 400):
  return jsonify({"ok": False, "error": message}), status


def safe_logo_path(path_value: str):
  path_text = clean_text(path_value)
  preferred = r"C:\Users\smith\Downloads\CNCPOWDER - Sbooks.jpg"
  if (not path_text or path_text.endswith("CNCPOWDER - qbooks_-02.jpg")) and clean_text(preferred):
    try:
      from pathlib import Path
      preferred_path = Path(preferred).expanduser().resolve()
      if preferred_path.exists() and preferred_path.is_file():
        return preferred_path
    except Exception:
      pass
  if not path_text:
    return None
  try:
    from pathlib import Path
    p = Path(path_text).expanduser().resolve()
  except Exception:
    return None
  return p if p.exists() and p.is_file() else None


def inline_logo_data_uri(path_value: str) -> str | None:
  logo_path = safe_logo_path(path_value)
  if not logo_path:
    return None
  return inline_file_data_uri(logo_path)


def inline_file_data_uri(path_value) -> str | None:
  file_path = Path(path_value)
  if not file_path.exists() or not file_path.is_file():
    return None
  mime, _ = mimetypes.guess_type(str(file_path))
  mime = mime or "image/jpeg"
  try:
    encoded = base64.b64encode(file_path.read_bytes()).decode("ascii")
  except Exception:
    return None
  return f"data:{mime};base64,{encoded}"


def sbooks_brand_data_uri() -> str | None:
  return inline_file_data_uri(SBOOKS_BRAND_ASSET)


def file_bytes_and_mime(path_value) -> tuple[bytes, str, str] | None:
  file_path = Path(path_value)
  if not file_path.exists() or not file_path.is_file():
    return None
  mime, _ = mimetypes.guess_type(str(file_path))
  mime = mime or "application/octet-stream"
  return file_path.read_bytes(), mime, file_path.name


def company_logo_file(settings: dict):
  return safe_logo_path(settings.get("company_logo_path", ""))


def browser_pdf_executable() -> str | None:
  candidates = [
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
  ]
  for path in candidates:
    if os.path.exists(path):
      return path
  return None


def jpeg_dimensions(image_bytes: bytes) -> tuple[int, int]:
  if len(image_bytes) < 4 or image_bytes[:2] != b"\xff\xd8":
    raise ValueError("Only JPEG logos are supported for PDF embedding right now")
  i = 2
  while i < len(image_bytes) - 9:
    if image_bytes[i] != 0xFF:
      i += 1
      continue
    marker = image_bytes[i + 1]
    i += 2
    if marker in {0xD8, 0xD9}:
      continue
    if i + 2 > len(image_bytes):
      break
    seg_len = struct.unpack(">H", image_bytes[i:i + 2])[0]
    if marker in {0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7, 0xC9, 0xCA, 0xCB, 0xCD, 0xCE, 0xCF}:
      if i + 7 > len(image_bytes):
        break
      height = struct.unpack(">H", image_bytes[i + 3:i + 5])[0]
      width = struct.unpack(">H", image_bytes[i + 5:i + 7])[0]
      return width, height
    i += seg_len
  raise ValueError("Could not determine JPEG size")


def build_pdf_bytes_with_jpeg(
  stream: bytes,
  jpeg_bytes: bytes | None = None,
  jpeg_width: int | None = None,
  jpeg_height: int | None = None,
  link_url: str | None = None,
  link_rect: tuple[float, float, float, float] | None = None,
) -> bytes:
  from generate_app_summary_pdf import PAGE_H, PAGE_W

  objects = []

  def add(obj: bytes) -> int:
    objects.append(obj)
    return len(objects)

  has_image = bool(jpeg_bytes and jpeg_width and jpeg_height)
  has_link = bool(link_url and link_rect)

  add(b"<< /Type /Catalog /Pages 2 0 R >>")
  add(b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>")
  if has_image and has_link:
    page_obj = (
      f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {PAGE_W} {PAGE_H}] "
      f"/Resources << /Font << /F1 5 0 R /F2 6 0 R >> /XObject << /Im1 7 0 R >> >> "
      f"/Annots [8 0 R] /Contents 4 0 R >>"
    ).encode("ascii")
  elif has_image:
    page_obj = (
      f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {PAGE_W} {PAGE_H}] "
      f"/Resources << /Font << /F1 5 0 R /F2 6 0 R >> /XObject << /Im1 7 0 R >> >> "
      f"/Contents 4 0 R >>"
    ).encode("ascii")
  else:
    page_obj = (
      f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 {PAGE_W} {PAGE_H}] "
      f"/Resources << /Font << /F1 5 0 R /F2 6 0 R >> >> /Contents 4 0 R >>"
    ).encode("ascii")
  add(page_obj)
  add(f"<< /Length {len(stream)} >>\nstream\n".encode("ascii") + stream + b"\nendstream")
  add(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>")
  add(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold >>")
  if has_image:
    add(
      f"<< /Type /XObject /Subtype /Image /Width {jpeg_width} /Height {jpeg_height} "
      f"/ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /DCTDecode /Length {len(jpeg_bytes)} >>\nstream\n".encode("ascii")
      + jpeg_bytes
      + b"\nendstream"
    )
  if has_link:
    x1, y1, x2, y2 = link_rect
    safe_url = str(link_url).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    add(
      f"<< /Type /Annot /Subtype /Link /Border [0 0 0] /Rect [{x1:.2f} {y1:.2f} {x2:.2f} {y2:.2f}] "
      f"/A << /Type /Action /S /URI /URI ({safe_url}) >> >>".encode("latin-1", "replace")
    )

  out = bytearray(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
  offsets = [0]
  for i, obj in enumerate(objects, start=1):
    offsets.append(len(out))
    out.extend(f"{i} 0 obj\n".encode("ascii"))
    out.extend(obj)
    out.extend(b"\nendobj\n")

  xref_start = len(out)
  out.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
  out.extend(b"0000000000 65535 f \n")
  for off in offsets[1:]:
    out.extend(f"{off:010d} 00000 n \n".encode("ascii"))
  out.extend(
    f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF\n".encode("ascii")
  )
  return bytes(out)


def business_settings_dict(conn) -> dict:
  rows = conn.execute("SELECT key, value FROM business_settings").fetchall()
  settings = BUSINESS_SETTINGS_DEFAULTS.copy()
  for row in rows:
    settings[row["key"]] = row["value"]
  return settings


def save_business_settings(conn, payload: dict) -> dict:
  settings = business_settings_dict(conn)
  for key, default in BUSINESS_SETTINGS_DEFAULTS.items():
    if key in payload:
      settings[key] = str(payload.get(key, default) or "")
      conn.execute(
        "INSERT OR REPLACE INTO business_settings(key, value) VALUES (?, ?)",
        (key, settings[key])
      )
  return settings


def next_document_number(conn, doc_type: str) -> str:
  settings = business_settings_dict(conn)
  prefix_key = "estimate_prefix" if doc_type == "estimate" else "invoice_prefix"
  next_key = "next_estimate_number" if doc_type == "estimate" else "next_invoice_number"
  prefix = settings.get(prefix_key, "EST-" if doc_type == "estimate" else "INV-")
  current = int(parse_float(settings.get(next_key, "1001"), 1001))
  number = f"{prefix}{current}"
  conn.execute(
    "INSERT OR REPLACE INTO business_settings(key, value) VALUES (?, ?)",
    (next_key, str(current + 1))
  )
  return number


def calculate_document_totals(lines: list[dict], tax_rate: float) -> tuple[float, float, float]:
  subtotal = 0.0
  taxable_subtotal = 0.0
  for line in lines:
    qty = parse_float(line.get("quantity"), 1.0)
    unit_price = parse_float(line.get("unit_price"), 0.0)
    line_total = round(qty * unit_price, 2)
    line["quantity"] = qty
    line["unit_price"] = unit_price
    line["line_total"] = line_total
    subtotal += line_total
    if parse_bool(line.get("taxable"), True):
      taxable_subtotal += line_total
  subtotal = round(subtotal, 2)
  tax_amount = round(taxable_subtotal * (parse_float(tax_rate) / 100.0), 2)
  total = round(subtotal + tax_amount, 2)
  return subtotal, tax_amount, total


def serialize_customer(row) -> dict:
  return {
    "id": int(row["id"]),
    "name": row["name"],
    "contact_name": row["contact_name"],
    "email": row["email"],
    "phone": row["phone"],
    "billing_address": row["billing_address"],
    "notes": row["notes"],
    "is_active": bool(row["is_active"]),
    "external_source": row["external_source"],
    "external_id": row["external_id"],
    "created_at": row["created_at"],
    "updated_at": row["updated_at"],
  }


def serialize_product(row) -> dict:
  return {
    "id": int(row["id"]),
    "name": row["name"],
    "description": row["description"],
    "sku": row["sku"],
    "default_unit_price": float(row["default_unit_price"] or 0),
    "taxable": bool(row["taxable"]),
    "is_active": bool(row["is_active"]),
    "external_source": row["external_source"],
    "external_id": row["external_id"],
    "created_at": row["created_at"],
    "updated_at": row["updated_at"],
  }


def fetch_customer(conn, customer_id: int):
  return conn.execute("SELECT * FROM customers WHERE id=?", (customer_id,)).fetchone()


def fetch_product(conn, product_id: int):
  return conn.execute("SELECT * FROM products WHERE id=?", (product_id,)).fetchone()


def fetch_document(conn, document_id: int):
  doc = conn.execute(
    """SELECT d.*, c.name AS customer_name, c.contact_name AS customer_contact_name,
              c.email AS customer_email, c.phone AS customer_phone,
              c.billing_address AS customer_billing_address
       FROM documents d
       JOIN customers c ON c.id = d.customer_id
       WHERE d.id=?""",
    (document_id,)
  ).fetchone()
  if not doc:
    return None
  lines = conn.execute(
    """SELECT dl.*, p.name AS product_name, p.sku AS product_sku
       FROM document_lines dl
       LEFT JOIN products p ON p.id = dl.product_id
       WHERE dl.document_id=?
       ORDER BY dl.sort_order, dl.id""",
    (document_id,)
  ).fetchall()
  return {
    "id": int(doc["id"]),
    "type": doc["type"],
    "number": doc["number"],
    "customer_id": int(doc["customer_id"]),
    "issue_date": doc["issue_date"],
    "due_date": doc["due_date"],
    "status": doc["status"],
    "subtotal": float(doc["subtotal"] or 0),
    "tax_rate": float(doc["tax_rate"] or 0),
    "tax_amount": float(doc["tax_amount"] or 0),
    "total": float(doc["total"] or 0),
    "notes": doc["notes"],
    "terms": doc["terms"],
    "imported": bool(doc["imported"]),
    "source_system": doc["source_system"],
    "source_id": doc["source_id"],
    "payment_url": doc["payment_url"],
    "last_sent_at": doc["last_sent_at"],
    "last_sent_to": doc["last_sent_to"],
    "last_email_error": doc["last_email_error"],
    "converted_from_document_id": doc["converted_from_document_id"],
    "created_at": doc["created_at"],
    "updated_at": doc["updated_at"],
    "customer": {
      "id": int(doc["customer_id"]),
      "name": doc["customer_name"],
      "contact_name": doc["customer_contact_name"],
      "email": doc["customer_email"],
      "phone": doc["customer_phone"],
      "billing_address": doc["customer_billing_address"],
    },
    "lines": [
      {
        "id": int(line["id"]),
        "product_id": line["product_id"],
        "description": line["description"],
        "quantity": float(line["quantity"] or 0),
        "unit_price": float(line["unit_price"] or 0),
        "taxable": bool(line["taxable"]),
        "line_total": float(line["line_total"] or 0),
        "sort_order": int(line["sort_order"] or 0),
        "product_name": line["product_name"],
        "product_sku": line["product_sku"],
      }
      for line in lines
    ],
  }


def build_document_payload(conn, payload: dict, existing: dict | None = None) -> dict:
  doc_type = clean_text(payload.get("type") or (existing or {}).get("type"))
  if doc_type not in {"estimate", "invoice"}:
    raise ValueError("Document type must be 'estimate' or 'invoice'")

  customer_id = int(payload.get("customer_id") or (existing or {}).get("customer_id") or 0)
  if not customer_id or not fetch_customer(conn, customer_id):
    raise ValueError("Valid customer is required")

  issue_date_value = payload.get("issue_date") or (existing or {}).get("issue_date") or date.today().isoformat()
  issue_date = iso(parse_date(issue_date_value))
  due_date_raw = payload.get("due_date", (existing or {}).get("due_date"))
  due_date = iso(parse_date(due_date_raw)) if due_date_raw else None

  lines_in = payload.get("lines")
  if lines_in is None and existing:
    lines_in = existing.get("lines", [])
  lines = []
  for idx, line in enumerate(lines_in or []):
    desc = clean_text(line.get("description"))
    if not desc and line.get("product_id"):
      prod = fetch_product(conn, int(line["product_id"]))
      if prod:
        desc = prod["description"] or prod["name"]
    if not desc:
      continue
    product_id = line.get("product_id")
    product_id = int(product_id) if product_id not in (None, "", "null") else None
    lines.append({
      "product_id": product_id,
      "description": desc,
      "quantity": parse_float(line.get("quantity"), 1.0),
      "unit_price": parse_float(line.get("unit_price"), 0.0),
      "taxable": bool(parse_bool(line.get("taxable"), True)),
      "sort_order": idx,
    })
  if not lines:
    raise ValueError("At least one line item is required")

  tax_rate = parse_float(payload.get("tax_rate"), parse_float((existing or {}).get("tax_rate"), parse_float(business_settings_dict(conn).get("default_tax_rate", "0"))))
  subtotal, tax_amount, total = calculate_document_totals(lines, tax_rate)
  status = clean_text(payload.get("status") or (existing or {}).get("status") or ("draft" if doc_type == "estimate" else "draft"))
  if doc_type == "estimate":
    status = "draft"
  elif status not in {"draft", "open", "paid"}:
    status = "draft"

  number = clean_text(payload.get("number") or (existing or {}).get("number"))
  if not number:
    number = next_document_number(conn, doc_type)

  return {
    "type": doc_type,
    "number": number,
    "customer_id": customer_id,
    "issue_date": issue_date,
    "due_date": due_date,
    "status": status,
    "subtotal": subtotal,
    "tax_rate": tax_rate,
    "tax_amount": tax_amount,
    "total": total,
    "notes": clean_text(payload.get("notes") or (existing or {}).get("notes")),
    "terms": clean_text(payload.get("terms") or (existing or {}).get("terms") or business_settings_dict(conn).get("default_terms", "")),
    "imported": int(bool(payload.get("imported", (existing or {}).get("imported", False)))),
    "source_system": clean_text(payload.get("source_system") or (existing or {}).get("source_system")) or None,
    "source_id": clean_text(payload.get("source_id") or (existing or {}).get("source_id")) or None,
    "converted_from_document_id": payload.get("converted_from_document_id", (existing or {}).get("converted_from_document_id")),
    "lines": lines,
  }


def save_document(conn, payload: dict, document_id: int | None = None) -> int:
  existing = fetch_document(conn, document_id) if document_id else None
  doc = build_document_payload(conn, payload, existing)
  if document_id:
    conn.execute(
      """UPDATE documents
         SET type=?, number=?, customer_id=?, issue_date=?, due_date=?, status=?,
             subtotal=?, tax_rate=?, tax_amount=?, total=?, notes=?, terms=?,
             imported=?, source_system=?, source_id=?, converted_from_document_id=?, updated_at=?
         WHERE id=?""",
      (
        doc["type"], doc["number"], doc["customer_id"], doc["issue_date"], doc["due_date"], doc["status"],
        doc["subtotal"], doc["tax_rate"], doc["tax_amount"], doc["total"], doc["notes"], doc["terms"],
        doc["imported"], doc["source_system"], doc["source_id"], doc["converted_from_document_id"], now_iso(),
        document_id,
      )
    )
    conn.execute("DELETE FROM document_lines WHERE document_id=?", (document_id,))
    target_id = document_id
  else:
    cur = conn.execute(
      """INSERT INTO documents
         (type, number, customer_id, issue_date, due_date, status, subtotal, tax_rate, tax_amount, total,
          notes, terms, imported, source_system, source_id, converted_from_document_id, created_at, updated_at)
         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
      (
        doc["type"], doc["number"], doc["customer_id"], doc["issue_date"], doc["due_date"], doc["status"],
        doc["subtotal"], doc["tax_rate"], doc["tax_amount"], doc["total"], doc["notes"], doc["terms"],
        doc["imported"], doc["source_system"], doc["source_id"], doc["converted_from_document_id"],
        now_iso(), now_iso(),
      )
    )
    target_id = cur.lastrowid

  for line in doc["lines"]:
    conn.execute(
      """INSERT INTO document_lines
         (document_id, product_id, description, quantity, unit_price, taxable, line_total, sort_order)
         VALUES (?,?,?,?,?,?,?,?)""",
      (
        target_id,
        line["product_id"],
        line["description"],
        line["quantity"],
        line["unit_price"],
        int(bool(line["taxable"])),
        line["line_total"],
        int(line["sort_order"]),
      )
    )
  return int(target_id)


def upsert_customer(conn, payload: dict) -> int:
  source = clean_text(payload.get("external_source")) or None
  external_id = clean_text(payload.get("external_id")) or None
  name = clean_text(payload.get("name"))
  if not name:
    raise ValueError("Customer name is required")
  existing = None
  if source and external_id:
    existing = conn.execute(
      "SELECT id FROM customers WHERE external_source=? AND external_id=?",
      (source, external_id)
    ).fetchone()
  if not existing:
    existing = conn.execute(
      "SELECT id FROM customers WHERE lower(name)=lower(?)",
      (name,)
    ).fetchone()
  fields = (
    name,
    clean_text(payload.get("contact_name")) or None,
    clean_text(payload.get("email")) or None,
    clean_text(payload.get("phone")) or None,
    clean_text(payload.get("billing_address")) or None,
    clean_text(payload.get("notes")) or None,
    int(bool(payload.get("is_active", True))),
    source,
    external_id,
    now_iso(),
  )
  if existing:
    conn.execute(
      """UPDATE customers
         SET name=?, contact_name=?, email=?, phone=?, billing_address=?, notes=?, is_active=?,
             external_source=?, external_id=?, updated_at=?
         WHERE id=?""",
      (*fields, int(existing["id"]))
    )
    return int(existing["id"])
  cur = conn.execute(
    """INSERT INTO customers
       (name, contact_name, email, phone, billing_address, notes, is_active,
        external_source, external_id, created_at, updated_at)
       VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
    (*fields[:-1], now_iso(), now_iso())
  )
  return int(cur.lastrowid)


def upsert_product(conn, payload: dict) -> int:
  source = clean_text(payload.get("external_source")) or None
  external_id = clean_text(payload.get("external_id")) or None
  name = clean_text(payload.get("name"))
  sku = clean_text(payload.get("sku")) or None
  if not name:
    raise ValueError("Product name is required")
  existing = None
  if source and external_id:
    existing = conn.execute(
      "SELECT id FROM products WHERE external_source=? AND external_id=?",
      (source, external_id)
    ).fetchone()
  if not existing and sku:
    existing = conn.execute(
      "SELECT id FROM products WHERE lower(sku)=lower(?)",
      (sku,)
    ).fetchone()
  if not existing:
    existing = conn.execute(
      "SELECT id FROM products WHERE lower(name)=lower(?)",
      (name,)
    ).fetchone()
  fields = (
    name,
    clean_text(payload.get("description")) or None,
    sku,
    parse_float(payload.get("default_unit_price"), 0.0),
    int(bool(payload.get("taxable", True))),
    int(bool(payload.get("is_active", True))),
    source,
    external_id,
    now_iso(),
  )
  if existing:
    conn.execute(
      """UPDATE products
         SET name=?, description=?, sku=?, default_unit_price=?, taxable=?, is_active=?,
             external_source=?, external_id=?, updated_at=?
         WHERE id=?""",
      (*fields, int(existing["id"]))
    )
    return int(existing["id"])
  cur = conn.execute(
    """INSERT INTO products
       (name, description, sku, default_unit_price, taxable, is_active,
        external_source, external_id, created_at, updated_at)
       VALUES (?,?,?,?,?,?,?,?,?,?)""",
    (*fields[:-1], now_iso(), now_iso())
  )
  return int(cur.lastrowid)


def _read_table_bytes(filename: str, payload: bytes) -> list[dict]:
  lower = filename.lower()
  if lower.endswith(".csv"):
    text = payload.decode("utf-8-sig")
    rows = []
    for row in csv.DictReader(io.StringIO(text)):
      rows.append({normalize_header(k): clean_text(v) for k, v in row.items()})
    return rows
  if lower.endswith(".xlsx"):
    try:
      from openpyxl import load_workbook
      wb = load_workbook(io.BytesIO(payload), data_only=True)
      sheet = wb.active
      rows = list(sheet.iter_rows(values_only=True))
    except Exception:
      ns = {
        "a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
      }
      root_zip = zipfile.ZipFile(io.BytesIO(payload))
      shared = []
      if "xl/sharedStrings.xml" in root_zip.namelist():
        import xml.etree.ElementTree as ET
        shared_root = ET.fromstring(root_zip.read("xl/sharedStrings.xml"))
        for si in shared_root.findall("a:si", ns):
          shared.append("".join(t.text or "" for t in si.findall(".//a:t", ns)))
      import xml.etree.ElementTree as ET
      wb_root = ET.fromstring(root_zip.read("xl/workbook.xml"))
      rels_root = ET.fromstring(root_zip.read("xl/_rels/workbook.xml.rels"))
      rel_map = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_root.findall("pr:Relationship", ns)}
      first_sheet = wb_root.findall("a:sheets/a:sheet", ns)[0]
      rel_id = first_sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
      target = rel_map[rel_id]
      path = "xl/" + target.lstrip("/") if not target.startswith("xl/") else target
      sheet_root = ET.fromstring(root_zip.read(path))
      rows = []
      for row in sheet_root.findall("a:sheetData/a:row", ns):
        values = []
        for cell in row.findall("a:c", ns):
          cell_type = cell.attrib.get("t")
          value_node = cell.find("a:v", ns)
          if cell_type == "s" and value_node is not None and value_node.text is not None:
            idx = int(value_node.text)
            values.append(shared[idx] if idx < len(shared) else "")
          elif cell_type == "inlineStr":
            values.append("".join(t.text or "" for t in cell.findall(".//a:t", ns)))
          elif value_node is not None and value_node.text is not None:
            values.append(value_node.text)
          else:
            values.append("")
        rows.append(values)
    if not rows:
      return []
    header_idx = 0
    for idx, candidate in enumerate(rows[:10]):
      normalized = [normalize_header(v) for v in candidate if normalize_header(v)]
      if len(normalized) >= 2 and not normalized[0].startswith("cnc_powder"):
        header_idx = idx
        break
    headers = [normalize_header(v) for v in rows[header_idx]]
    out = []
    for values in rows[header_idx + 1:]:
      if not any(v not in (None, "") for v in values):
        continue
      out.append({
        headers[idx]: clean_text(val)
        for idx, val in enumerate(values)
        if idx < len(headers) and headers[idx]
      })
    return out
  raise ValueError("Only CSV and XLSX files are supported")


def read_table_file(file_storage) -> list[dict]:
  if not file_storage or not file_storage.filename:
    return []
  return _read_table_bytes(file_storage.filename, file_storage.read())


def normalize_customer_rows(rows: list[dict]) -> tuple[list[dict], list[str], list[str]]:
  items, warnings, errors = [], [], []
  for idx, row in enumerate(rows, start=2):
    name = clean_text(pick_value(row, "display_name", "customer_name", "customer", "name", "company", "company_name", "full_name"))
    if looks_like_report_footer(name):
      continue
    if not name:
      errors.append(f"Customers row {idx}: missing customer name")
      continue
    billing_address = clean_text(pick_value(row, "billing_address", "bill_address", "billingaddress"))
    if not billing_address:
      address_parts = [
        pick_value(row, "billing_address_line_1", "bill_addr1", "address", "address1"),
        pick_value(row, "billing_address_line_2", "bill_addr2", "address2"),
        pick_value(row, "billing_city", "city"),
        pick_value(row, "billing_state", "state"),
        pick_value(row, "billing_postal_code", "zip", "postal_code"),
      ]
      billing_address = ", ".join([clean_text(part) for part in address_parts if clean_text(part)])
    items.append({
      "name": name,
      "contact_name": clean_text(pick_value(row, "contact", "contact_name", "attention", "attn")),
      "email": clean_text(pick_value(row, "email", "email_address", "primary_email")),
      "phone": clean_text(pick_value(row, "phone", "phone_number", "main_phone", "mobile")),
      "billing_address": billing_address,
      "notes": clean_text(pick_value(row, "notes", "memo")),
      "is_active": parse_bool(pick_value(row, "active", "is_active"), True),
      "external_source": "quickbooks",
      "external_id": clean_text(pick_value(row, "list_id", "customer_id", "id", "quickbooks_id")),
    })
  return items, warnings, errors


def normalize_product_rows(rows: list[dict]) -> tuple[list[dict], list[str], list[str]]:
  items, warnings, errors = [], [], []
  for idx, row in enumerate(rows, start=2):
    name = clean_text(pick_value(
      row,
      "product_service",
      "product_service_name",
      "product_service_full_name",
      "item_name",
      "item",
      "name",
      "full_name",
    ))
    if looks_like_report_footer(name):
      continue
    if not name:
      errors.append(f"Products row {idx}: missing product/service name")
      continue
    items.append({
      "name": name,
      "description": clean_text(pick_value(
        row,
        "description",
        "memo_description",
        "sales_description",
        "purchase_description",
      )) or name,
      "sku": clean_text(pick_value(row, "sku", "item_code", "code")),
      "default_unit_price": parse_float(pick_value(
        row,
        "sales_price_rate",
        "sales_price",
        "rate",
        "unit_price",
        "price",
      ), 0.0),
      "taxable": parse_bool(pick_value(row, "taxable", "is_taxable", "tax"), True),
      "is_active": parse_bool(pick_value(row, "active", "is_active"), True),
      "external_source": "quickbooks",
      "external_id": clean_text(pick_value(row, "list_id", "product_id", "service_id", "id", "quickbooks_id")),
    })
  return items, warnings, errors


def normalize_invoice_rows(rows: list[dict]) -> tuple[list[dict], list[str], list[str]]:
  grouped: dict[str, dict] = {}
  warnings, errors = [], []
  for idx, row in enumerate(rows, start=2):
    invoice_number = clean_text(pick_value(row, "invoice_number", "invoice_no", "invoice_num", "doc_number", "ref_number", "num", "no"))
    source_id = clean_text(pick_value(row, "txn_id", "transaction_id", "invoice_id", "id"))
    if looks_like_report_footer(invoice_number) or looks_like_report_footer(source_id):
      continue
    if not invoice_number and not source_id:
      errors.append(f"Invoices row {idx}: missing invoice number/id")
      continue
    key = source_id or invoice_number
    customer_name = clean_text(pick_value(row, "customer", "customer_name", "customer_full_name", "name"))
    if not customer_name:
      customer_name = "Imported Customer"
      warnings.append(f"Invoices row {idx}: missing customer name, using placeholder")
    doc = grouped.setdefault(key, {
      "type": "invoice",
      "number": invoice_number or f"QB-{key}",
      "customer_name": customer_name,
      "issue_date": clean_text(pick_value(row, "invoice_date", "txn_date", "date")) or date.today().isoformat(),
      "due_date": clean_text(pick_value(row, "due_date", "due")),
      "status": "open",
      "tax_rate": parse_float(pick_value(row, "tax_rate", "sales_tax_rate"), 0.0),
      "notes": clean_text(pick_value(row, "memo", "message", "notes")),
      "source_system": "quickbooks",
      "source_id": source_id or invoice_number,
      "imported": True,
      "lines": [],
    })
    if not doc["due_date"]:
      doc["due_date"] = clean_text(pick_value(row, "due_date", "due"))
    if not doc["notes"]:
      doc["notes"] = clean_text(pick_value(row, "memo", "message", "notes"))
    line_description = clean_text(pick_value(row, "description", "item_description", "product_service", "product_service_name", "item", "item_name"))
    quantity = parse_float(pick_value(row, "qty", "quantity"), 1.0)
    unit_price = parse_float(pick_value(row, "rate", "unit_price", "price", "sales_price"), 0.0)
    amount = parse_float(pick_value(row, "line_amount", "amount"), quantity * unit_price)
    product_name = clean_text(pick_value(row, "product_service", "product_service_name", "item", "item_name"))
    if not line_description and amount == 0 and clean_text(pick_value(row, "total", "invoice_total")):
      continue
    if not line_description:
      line_description = product_name or "Imported line item"
    if quantity == 0:
      quantity = 1.0
    if unit_price == 0 and quantity:
      unit_price = round(amount / quantity, 2) if amount else 0.0
    doc["lines"].append({
      "product_name": product_name,
      "description": line_description,
      "quantity": quantity,
      "unit_price": unit_price,
      "taxable": parse_bool(pick_value(row, "taxable", "is_taxable", "tax"), True),
    })

  documents = []
  for key, doc in grouped.items():
    if not doc["lines"]:
      documents.append({
        **doc,
        "lines": [{
          "product_name": "",
          "description": "Imported invoice total",
          "quantity": 1.0,
          "unit_price": parse_float(0),
          "taxable": True,
        }]
      })
      warnings.append(f"Invoice {doc['number']}: no line items found, created placeholder line")
    else:
      documents.append(doc)
  return documents, warnings, errors


def normalize_invoice_rows_from_general_ledger(rows: list[dict]) -> tuple[list[dict], list[str], list[str]]:
  documents, warnings, errors = [], [], []
  seen_numbers = set()
  for idx, row in enumerate(rows, start=2):
    tx_type = clean_text(pick_value(row, "transaction_type"))
    if looks_like_report_footer(tx_type):
      continue
    if tx_type.lower() != "invoice":
      continue
    account = clean_text(pick_value(row, "account"))
    if account and "accounts receivable" not in account.lower():
      continue
    number = clean_text(pick_value(row, "num", "number"))
    customer_name = clean_text(pick_value(row, "name")) or "Imported Customer"
    amount = parse_float(pick_value(row, "debit", "credit"), 0.0)
    if not number:
      warnings.append(f"General ledger row {idx}: missing invoice number, skipped")
      continue
    if number in seen_numbers:
      continue
    seen_numbers.add(number)
    documents.append({
      "type": "invoice",
      "number": number.split(".")[0] if number.endswith(".0") else number,
      "customer_name": customer_name,
      "issue_date": clean_text(pick_value(row, "date")) or date.today().isoformat(),
      "due_date": "",
      "status": "open",
      "tax_rate": 0.0,
      "notes": clean_text(pick_value(row, "memo_description")),
      "source_system": "quickbooks",
      "source_id": clean_text(pick_value(row, "num", "number")) or number,
      "imported": True,
      "lines": [{
        "product_name": "",
        "description": clean_text(pick_value(row, "memo_description")) or "Imported invoice total",
        "quantity": 1.0,
        "unit_price": amount,
        "taxable": True,
      }],
    })
  return documents, warnings, errors


def preview_quickbooks_bundle(file_storage) -> dict:
  warnings = []
  errors = []
  customers = []
  products = []
  documents = []
  if not file_storage or not file_storage.filename:
    return {"customers": customers, "products": products, "documents": documents, "warnings": warnings, "errors": errors}
  if not file_storage.filename.lower().endswith(".zip"):
    raise ValueError("QuickBooks backup upload must be a ZIP file")
  with zipfile.ZipFile(io.BytesIO(file_storage.read())) as archive:
    entries = {name.split("/")[-1].lower(): name for name in archive.namelist() if name.lower().endswith((".csv", ".xlsx"))}
    if "customers.xlsx" in entries or "customers.csv" in entries:
      entry = entries.get("customers.xlsx") or entries.get("customers.csv")
      rows = _read_table_bytes(entry, archive.read(entry))
      customers, c_warn, c_err = normalize_customer_rows(rows)
      warnings.extend(c_warn)
      errors.extend(c_err)
    else:
      warnings.append("Customers export not found in backup zip")

    product_entry = entries.get("products.xlsx") or entries.get("products.csv") or entries.get("services.xlsx") or entries.get("items.xlsx")
    if product_entry:
      rows = _read_table_bytes(product_entry, archive.read(product_entry))
      products, p_warn, p_err = normalize_product_rows(rows)
      warnings.extend(p_warn)
      errors.extend(p_err)
    else:
      warnings.append("Products/services export not found in backup zip")

    invoice_entry = entries.get("invoices.xlsx") or entries.get("invoices.csv")
    if invoice_entry:
      rows = _read_table_bytes(invoice_entry, archive.read(invoice_entry))
      documents, d_warn, d_err = normalize_invoice_rows(rows)
      warnings.extend(d_warn)
      errors.extend(d_err)
    else:
      gl_entry = entries.get("general_ledger.xlsx") or entries.get("general_ledger.csv")
      if gl_entry:
        rows = _read_table_bytes(gl_entry, archive.read(gl_entry))
        documents, d_warn, d_err = normalize_invoice_rows_from_general_ledger(rows)
        warnings.extend(d_warn)
        errors.extend(d_err)
      else:
        warnings.append("Invoice export not found in backup zip")
  return {
    "customers": customers,
    "products": products,
    "documents": documents,
    "warnings": warnings,
    "errors": errors,
  }


def render_document_pdf(document: dict, settings: dict) -> bytes:
  from textwrap import wrap
  from generate_app_summary_pdf import Canvas, PAGE_H, PAGE_W

  left = 42
  right = PAGE_W - 42
  canvas = Canvas()
  website = clean_text(settings.get("company_website")) or "https://cncpowder.com"
  jpeg_bytes = None
  jpeg_width = None
  jpeg_height = None
  link_rect = None
  logo_path = safe_logo_path(settings.get("company_logo_path", ""))
  if logo_path and logo_path.suffix.lower() in {".jpg", ".jpeg"}:
    try:
      jpeg_bytes = logo_path.read_bytes()
      jpeg_width, jpeg_height = jpeg_dimensions(jpeg_bytes)
      logo_box_w = 250
      logo_box_h = 68
      logo_box_x = right - logo_box_w - 18
      logo_box_y = PAGE_H - 182
      scale = min(logo_box_w / jpeg_width, logo_box_h / jpeg_height)
      draw_w = jpeg_width * scale
      draw_h = jpeg_height * scale
      draw_x = logo_box_x + (logo_box_w - draw_w)
      draw_y = logo_box_y + (logo_box_h - draw_h) / 2
      canvas.ops.append(f"q {draw_w:.2f} 0 0 {draw_h:.2f} {draw_x:.2f} {draw_y:.2f} cm /Im1 Do Q")
      canvas.text(min(right - 34, draw_x + draw_w + 4), draw_y + 3, 11, ".com", font="F2")
      link_rect = (logo_box_x, logo_box_y, right - 10, logo_box_y + logo_box_h)
    except Exception:
      jpeg_bytes = None
      jpeg_width = None
      jpeg_height = None
      link_rect = None
  y = PAGE_H - 130
  canvas.text(left, y, 22, settings.get("company_name", "My Business"), font="F2")
  y -= 20
  for line in [settings.get("company_address", ""), settings.get("company_phone", ""), settings.get("company_email", "")]:
    for subline in str(line or "").splitlines():
      if subline.strip():
        canvas.text(left + 4, y, 9, subline[:58])
        y -= 10

  y -= 14
  canvas.rect(left, y - 70, right - left, 70, fill_gray=0.94)
  canvas.text(left + 18, y - 24, 15, f"{document['type'].title()}", font="F2")
  canvas.text(left + 18, y - 43, 11, f"{document['number']}")
  canvas.text(right - 110, y - 42, 9, "AMOUNT DUE")
  canvas.text(right - 132, y - 62, 20, f"${document['total']:,.2f}", font="F2")
  y -= 86

  gap = 16
  card_w = (right - left - gap) / 2
  card_h = 90
  canvas.rect(left, y - card_h, card_w, card_h)
  canvas.rect(left + card_w + gap, y - card_h, card_w, card_h)
  canvas.text(left + 14, y - 18, 10, "BILL TO", font="F2")
  bill_y = y - 34
  for line in [document["customer"]["name"], document["customer"].get("contact_name"), document["customer"].get("email"), document["customer"].get("phone"), document["customer"].get("billing_address")]:
    for subline in str(line or "").splitlines():
      if subline.strip() and bill_y > y - card_h + 12:
        canvas.text(left + 14, bill_y, 9, subline[:42])
        bill_y -= 10

  meta_x = left + card_w + gap + 14
  canvas.text(meta_x, y - 18, 10, "DETAILS", font="F2")
  meta_y = y - 34
  for line in [
    f"Issue: {document['issue_date']}",
    f"Due: {document.get('due_date') or 'On receipt'}",
    f"Status: {str(document.get('status') or '').title()}",
    f"Tax rate: {document['tax_rate']:.2f}%",
  ]:
    canvas.text(meta_x, meta_y, 9, line[:44])
    meta_y -= 10

  y -= (card_h + 18)
  canvas.rect(left, y - 22, right - left, 22, fill_gray=0.95)
  canvas.text(left + 12, y - 14, 9, "DESCRIPTION", font="F2")
  canvas.text(right - 170, y - 14, 9, "QTY", font="F2")
  canvas.text(right - 110, y - 14, 9, "RATE", font="F2")
  canvas.text(right - 54, y - 14, 9, "AMOUNT", font="F2")
  y -= 34
  for line in document["lines"]:
    desc_lines = wrap(str(line["description"]), width=52, break_long_words=False, break_on_hyphens=False) or [""]
    canvas.text(left, y, 9, desc_lines[0])
    canvas.text(right - 165, y, 9, f"{float(line['quantity']):.2f}".rstrip("0").rstrip("."))
    canvas.text(right - 115, y, 9, f"${float(line['unit_price']):,.2f}")
    canvas.text(right - 55, y, 9, f"${float(line['line_total']):,.2f}")
    y -= 11
    for extra in desc_lines[1:]:
      canvas.text(left + 10, y, 9, extra[:56])
      y -= 11
    canvas.rule(left, y + 4, right, y + 4, width=0.4)
    y -= 6

  section_top = y - 2
  notes_x = left
  notes_w = 248
  terms_x = left + notes_w + 14
  terms_w = 142
  totals_x = right - 190

  if document.get("notes"):
    canvas.rect(notes_x, section_top - 70, notes_w, 70)
    canvas.text(notes_x + 12, section_top - 18, 10, "NOTES", font="F2")
    inner_y = section_top - 32
    for note_line in wrap(str(document["notes"]), width=42, break_long_words=False, break_on_hyphens=False)[:4]:
      canvas.text(notes_x + 12, inner_y, 9, note_line)
      inner_y -= 10
  if document.get("terms"):
    canvas.rect(terms_x, section_top - 70, terms_w, 70)
    canvas.text(terms_x + 12, section_top - 18, 10, "TERMS", font="F2")
    inner_y = section_top - 32
    for term_line in wrap(str(document["terms"]), width=20, break_long_words=False, break_on_hyphens=False)[:4]:
      canvas.text(terms_x + 12, inner_y, 9, term_line)
      inner_y -= 10

  canvas.rect(totals_x, section_top - 82, 190, 82)
  line_y = section_top - 18
  for label, value, bold in [
    ("Subtotal", document["subtotal"], False),
    (f"Tax ({document['tax_rate']:.2f}%)", document["tax_amount"], False),
    ("Total", document["total"], True),
  ]:
    canvas.text(totals_x + 12, line_y, 10 if bold else 9, label, font="F2" if bold else "F1")
    canvas.text(totals_x + 116, line_y, 10 if bold else 9, f"${value:,.2f}", font="F2" if bold else "F1")
    line_y -= 18
  return build_pdf_bytes_with_jpeg(
    canvas.render(),
    jpeg_bytes=jpeg_bytes,
    jpeg_width=jpeg_width,
    jpeg_height=jpeg_height,
    link_url=website,
    link_rect=link_rect,
  )


def render_document_pdf_bytes(document: dict, settings: dict) -> bytes:
  browser = browser_pdf_executable()
  if browser:
    html_doc = render_template(
      "document_print.html",
      title=document["number"],
      document=document,
      business=settings,
      logo_src=inline_logo_data_uri(settings.get("company_logo_path", "")),
    )
    try:
      work_tmp_root = os.path.join(os.getcwd(), "data", "_pdf_cache")
      os.makedirs(work_tmp_root, exist_ok=True)
      token = uuid.uuid4().hex
      html_path = os.path.join(work_tmp_root, f"{document['number']}-{token}.html")
      pdf_path = os.path.join(work_tmp_root, f"{document['number']}-{token}.pdf")
      with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_doc)
      subprocess.run(
        [
          browser,
          "--headless",
          "--disable-gpu",
          f"--user-data-dir={os.path.join(work_tmp_root, f'profile-{token}')}",
          "--allow-file-access-from-files",
          f"--print-to-pdf={pdf_path}",
          "--no-pdf-header-footer",
          Path(html_path).resolve().as_uri(),
        ],
        check=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        timeout=30,
      )
      with open(pdf_path, "rb") as f:
        return f.read()
    except Exception:
      pass
  return render_document_pdf(document, settings)


def app_base_url() -> str:
  return request.url_root.rstrip("/")


def build_payment_url(document: dict, settings: dict) -> str:
  base = clean_text(settings.get("invoice_payment_url_base"))
  if base:
    if "{id}" in base or "{number}" in base:
      return base.replace("{id}", str(document["id"])).replace("{number}", quote(str(document["number"])))
    return f"{base.rstrip('/')}/{document['id']}"
  return f"{app_base_url()}/pay/{document['id']}"


def ensure_document_payment_url(conn, document: dict, settings: dict, persist: bool = False) -> str:
  payment_url = build_payment_url(document, settings)
  if persist and payment_url != (document.get("payment_url") or ""):
    conn.execute(
      "UPDATE documents SET payment_url=?, updated_at=? WHERE id=?",
      (payment_url, now_iso(), document["id"])
    )
  document["payment_url"] = payment_url
  return payment_url


def build_invoice_email_draft(document: dict, settings: dict) -> dict:
  payment_url = build_payment_url(document, settings)
  recipient_name = document["customer"].get("contact_name") or document["customer"]["name"]
  intro_text = (
    f"Hi {recipient_name},\n\n"
    f"Your invoice {document['number']} from {settings.get('company_name') or APP_TITLE} is ready. "
    f"I've attached the PDF for your records."
  )
  due_line = (
    f"Due date: {document['due_date']}"
    if document.get("due_date")
    else "Due date: On receipt"
  )
  due_label = "Due date"
  due_value = document.get("due_date") or "On receipt"
  closing_text = (
    f"Questions? Reply to this email or call {settings.get('company_phone') or 'our office'}.\n\n"
    f"Thank you,\n{settings.get('company_name') or APP_TITLE}"
  )
  subject = f"Invoice {document['number']} from {settings.get('company_name') or APP_TITLE}"
  company_logo_path = company_logo_file(settings)
  html_body = render_template(
    "invoice_email.html",
    title=subject,
    document=document,
    business=settings,
    payment_url=payment_url,
    intro_html=html.escape(intro_text).replace("\n", "<br>"),
    closing_html=html.escape(closing_text).replace("\n", "<br>"),
    due_line=due_line,
    due_label=due_label,
    due_value=due_value,
    logo_src="cid:company-logo" if company_logo_path else None,
    sbooks_logo_src="cid:sbooks-logo",
  )
  preview_html = render_template(
    "invoice_email.html",
    title=subject,
    document=document,
    business=settings,
    payment_url=payment_url,
    intro_html=html.escape(intro_text).replace("\n", "<br>"),
    closing_html=html.escape(closing_text).replace("\n", "<br>"),
    due_line=due_line,
    due_label=due_label,
    due_value=due_value,
    logo_src=inline_logo_data_uri(settings.get("company_logo_path", "")),
    sbooks_logo_src=sbooks_brand_data_uri(),
  )
  text_body = "\n".join([
    intro_text,
    "",
    f"Amount due: ${document['total']:,.2f}",
    due_line,
    f"Pay online: {payment_url}",
    "",
    closing_text,
  ])
  return {
    "to": document["customer"].get("email") or "",
    "subject": subject,
    "html": html_body,
    "preview_html": preview_html,
    "text": text_body,
    "payment_url": payment_url,
    "inline_images": [
      {"cid": "sbooks-logo", "path": str(SBOOKS_BRAND_ASSET)},
      *([{"cid": "company-logo", "path": str(company_logo_path)}] if company_logo_path else []),
    ],
  }


def send_invoice_email_message(settings: dict, to_email: str, subject: str, html_body: str, text_body: str, pdf_name: str, pdf_bytes: bytes, inline_images: list[dict] | None = None):
  smtp_host = clean_text(settings.get("smtp_host"))
  smtp_port = int(parse_float(settings.get("smtp_port"), 587))
  smtp_username = clean_text(settings.get("smtp_username"))
  smtp_password = clean_text(settings.get("smtp_password"))
  company_email = clean_text(settings.get("company_email"))
  from_name = clean_text(settings.get("smtp_from_name")) or clean_text(settings.get("company_name")) or APP_TITLE
  from_email = company_email or smtp_username
  if not smtp_host:
    raise ValueError("SMTP host is required before sending invoice email")
  if not from_email:
    raise ValueError("Set company email before sending invoice email")
  if not clean_text(to_email):
    raise ValueError("Recipient email is required")

  msg = EmailMessage()
  msg["Subject"] = subject
  msg["From"] = formataddr((from_name, from_email))
  msg["To"] = clean_text(to_email)
  if company_email and company_email.lower() != clean_text(to_email).lower():
    msg["Reply-To"] = company_email
  msg.set_content(text_body or "Please see the attached invoice.")
  msg.add_alternative(html_body or "<p>Please see the attached invoice.</p>", subtype="html")
  html_part = msg.get_payload()[-1]
  for item in inline_images or []:
    file_info = file_bytes_and_mime(item.get("path"))
    if not file_info:
      continue
    file_bytes, mime, filename = file_info
    maintype, subtype = mime.split("/", 1) if "/" in mime else ("application", "octet-stream")
    html_part.add_related(
      file_bytes,
      maintype=maintype,
      subtype=subtype,
      cid=f"<{item['cid']}>",
      filename=filename,
      disposition="inline",
    )
  msg.add_attachment(pdf_bytes, maintype="application", subtype="pdf", filename=pdf_name)

  use_tls = parse_bool(settings.get("smtp_use_tls"), True)
  if use_tls:
    with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as server:
      server.ehlo()
      server.starttls(context=ssl.create_default_context())
      server.ehlo()
      if smtp_username:
        server.login(smtp_username, smtp_password)
      server.send_message(msg)
  else:
    with smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30, context=ssl.create_default_context()) as server:
      if smtp_username:
        server.login(smtp_username, smtp_password)
      server.send_message(msg)


def _archived_period_starts(conn) -> set[str]:
  try:
    rows = conn.execute("SELECT start_date FROM pay_period_archives").fetchall()
    return {r["start_date"] for r in rows}
  except Exception:
    return set()

def bucket_from_due_day(due_day: Optional[int]) -> str:
  if due_day is None:
    return "ONE_OFF"
  return "FIRST_OF_MONTH" if 1 <= due_day <= FIRST_OF_MONTH_MAX_DAY else "MID_MONTH"

def prev_month(y: int, m: int) -> Tuple[int, int]:
  if m == 1:
    return (y - 1, 12)
  return (y, m - 1)

def first_last_paycheck_ids(paychecks: List[dict], year: int, month: int) -> Tuple[Optional[int], Optional[int]]:
  in_month = [p for p in paychecks if parse_date(p["date"]).year == year and parse_date(p["date"]).month == month]
  if not in_month:
    return None, None
  in_month.sort(key=lambda p: p["date"])
  return in_month[0]["id"], in_month[-1]["id"]

def assign_funding_paycheck_id(tx_year: int, tx_month: int, bucket: str, paychecks: List[dict]) -> Optional[int]:
  if bucket == "ONE_OFF":
    return None
  if bucket == "FIRST_OF_MONTH":
    py, pm = prev_month(tx_year, tx_month)
    _, last_id = first_last_paycheck_ids(paychecks, py, pm)
    return last_id
  if bucket == "MID_MONTH":
    first_id, _ = first_last_paycheck_ids(paychecks, tx_year, tx_month)
    return first_id
  return None

def daterange(start: date, end: date):
  d = start
  while d <= end:
    yield d
    d += timedelta(days=1)


# Card definitions are stored in cc_snapshots as rows with this special date.
# This allows safe, ALTER-only migrations while keeping card metadata (name/url/due_day)
# separate from per-paycheck balances.
CC_CARDDEF_DATE = "0000-00-00"


def cc_latest_by_name(conn, as_of_iso: str, side: str = "personal"):
  """
  Returns latest snapshot per card name with snapshot_date <= as_of_iso for a given side.
  Includes snapshot id + url for edit/delete.
  """
  rows = conn.execute(
    """SELECT cs.id, cs.name, cs.balance, cs.snapshot_date, cs.url
       FROM cc_snapshots cs
       JOIN (
         SELECT name, MAX(snapshot_date) AS snapshot_date
         FROM cc_snapshots
         WHERE snapshot_date <= ? AND side = ?
         GROUP BY name
       ) m
       ON cs.name = m.name AND cs.snapshot_date = m.snapshot_date
       WHERE cs.side = ?
       ORDER BY cs.name""",
    (as_of_iso, side, side)
  ).fetchall()

  cards = []
  total = 0.0
  for r in rows:
    bal = float(r["balance"])
    total += bal
    cards.append({
      "id": int(r["id"]),
      "name": r["name"],
      "balance": bal,
      "snapshot_date": r["snapshot_date"],
      "url": r["url"]
    })
  return cards, round(total, 2)


def cc_card_names(conn, side: str):
  """Return the authoritative card list for a side.

  Card-definition rows (snapshot_date == CC_CARDDEF_DATE) are the source of
  truth for which cards exist. This prevents old/renamed cards from showing up
  as "ghost" duplicates in later pay periods.

  If no card-definition rows exist yet, fall back to distinct names from
  existing snapshots.
  """
  defs = conn.execute(
    """SELECT name FROM cc_snapshots
         WHERE side = ? AND snapshot_date = ?
         ORDER BY name""",
    (side, CC_CARDDEF_DATE)
  ).fetchall()
  if defs:
    return [r["name"] for r in defs]

  rows = conn.execute(
    "SELECT DISTINCT name FROM cc_snapshots WHERE side = ? ORDER BY name",
    (side,)
  ).fetchall()
  return [r["name"] for r in rows]

def cc_latest_url(conn, name: str, as_of_iso: str, side: str):
  row = conn.execute(
    """SELECT url FROM cc_snapshots
         WHERE name = ? AND side = ? AND snapshot_date <= ? AND url IS NOT NULL AND url != ''
         ORDER BY snapshot_date DESC
         LIMIT 1""",
    (name, side, as_of_iso)
  ).fetchone()
  return row["url"] if row else None

def cc_latest_due_day(conn, name: str, as_of_iso: str, side: str):
  """Two-digit due day-of-month ('01'..'31') card metadata."""
  row = conn.execute(
    """SELECT due_day FROM cc_snapshots
         WHERE name = ? AND side = ? AND snapshot_date <= ? AND due_day IS NOT NULL AND due_day != ''
         ORDER BY snapshot_date DESC
         LIMIT 1""",
    (name, side, as_of_iso)
  ).fetchone()
  return row["due_day"] if row else None


def cc_card_defs(conn, side: str):
  """Return card definitions (name, url, due_day) from CC_CARDDEF_DATE rows."""
  rows = conn.execute(
    """SELECT id, name, url, due_day
         FROM cc_snapshots
         WHERE side = ? AND snapshot_date = ?
         ORDER BY name""",
    (side, CC_CARDDEF_DATE)
  ).fetchall()
  cards = []
  for r in rows:
    cards.append({
      "id": int(r["id"]),
      "name": r["name"],
      "url": r["url"],
      "due_day": r["due_day"]
    })
  return cards

def cc_cards_for_paycheck(conn, payday_iso: str, side: str = "personal"):
  """
  Returns one entry per card name for this side, for the *exact* payday snapshot_date.

  - If a card has a snapshot row exactly on payday_iso, uses it (id present).
  - If missing for this payday, returns a zero-balance placeholder (id=None),
    keeping the most recent known url (if any) so the Pay link still works.

  This prevents later pay periods from 'reusing' earlier snapshot rows and keeps
  balances independent per pay period.
  """
  names = cc_card_names(conn, side)
  cards = []
  total = 0.0
  for nm in names:
    r = conn.execute(
      """SELECT id, name, balance, snapshot_date, url, pay_status
           FROM cc_snapshots
           WHERE side = ? AND name = ? AND snapshot_date = ?
           LIMIT 1""",
      (side, nm, payday_iso)
    ).fetchone()

    due_day = cc_latest_due_day(conn, nm, payday_iso, side)

    if r:
      bal = float(r["balance"])
      total += bal
      url = r["url"] or cc_latest_url(conn, nm, payday_iso, side)
      cards.append({
        "id": int(r["id"]),
        "name": r["name"],
        "balance": bal,
        "snapshot_date": r["snapshot_date"],
        "url": url,
        "due_day": due_day,
        "pay_status": r["pay_status"]
      })
    else:
      url = cc_latest_url(conn, nm, payday_iso, side)
      cards.append({
        "id": None,
        "name": nm,
        "balance": 0.0,
        "snapshot_date": payday_iso,
        "url": url,
        "due_day": due_day,
        "pay_status": None
      })
  return cards, round(total, 2)


app = Flask(__name__, static_folder="static", template_folder="templates")

@app.get("/")
def index():
  return render_template("index.html", title=APP_TITLE)


@app.get("/business")
def business_index():
  return render_template("business.html", title="S-Books")

@app.get("/static/<path:path>")
def static_proxy(path):
  return send_from_directory("static", path)

@app.post("/api/setup")
def setup():
  payload = request.get_json(force=True)
  anchor_date = parse_date(payload["anchor_date"])
  paycheck_amount = float(payload.get("paycheck_amount", 0))
  paycheck_account_id = int(payload.get("paycheck_account_id", 1))
  horizon_days = int(payload.get("horizon_days", 365))
  anchor_balances = payload.get("anchor_balances", [])

  end_date = anchor_date + timedelta(days=horizon_days)

  with db() as conn:
    conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", ("anchor_date", iso(anchor_date)))
    conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", ("horizon_days", str(horizon_days)))
    conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", ("paycheck_amount", str(paycheck_amount)))
    conn.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", ("paycheck_account_id", str(paycheck_account_id)))

    for item in anchor_balances:
      conn.execute(
        "INSERT OR REPLACE INTO account_anchors(account_id, anchor_date, anchor_balance) VALUES (?,?,?)",
        (int(item["account_id"]), iso(anchor_date), float(item["anchor_balance"]))
      )

    conn.execute("DELETE FROM paychecks WHERE account_id=? AND date>=?", (paycheck_account_id, iso(anchor_date)))
    d = anchor_date
    while d <= end_date:
      conn.execute(
        "INSERT OR IGNORE INTO paychecks(account_id,date,amount,is_primary) VALUES(?,?,?,1)",
        (paycheck_account_id, iso(d), paycheck_amount)
      )
      d += timedelta(days=14)

  return jsonify({"ok": True})

@app.get("/api/settings")
def get_settings():
  with db() as conn:
    rows = conn.execute("SELECT key,value FROM settings").fetchall()
    settings = {r["key"]: r["value"] for r in rows}
  return jsonify(settings)

@app.get("/api/accounts")
def list_accounts():
  with db() as conn:
    accounts = [dict(r) for r in conn.execute("SELECT * FROM accounts ORDER BY id").fetchall()]
  return jsonify(accounts)


@app.get("/api/business_settings")
def get_business_settings():
  with db() as conn:
    return jsonify(business_settings_dict(conn))


@app.get("/api/business_logo")
def get_business_logo():
  with db() as conn:
    settings = business_settings_dict(conn)
  logo_path = safe_logo_path(settings.get("company_logo_path", ""))
  if not logo_path:
    return json_error("Business logo not configured", 404)
  return send_file(str(logo_path))


@app.put("/api/business_settings")
def update_business_settings():
  payload = request.get_json(force=True) or {}
  with db() as conn:
    settings = save_business_settings(conn, payload)
  return jsonify({"ok": True, "settings": settings})


@app.get("/api/customers")
def list_customers():
  active_only = request.args.get("active_only") == "1"
  q = "SELECT * FROM customers"
  if active_only:
    q += " WHERE is_active=1"
  q += " ORDER BY lower(name)"
  with db() as conn:
    rows = conn.execute(q).fetchall()
  return jsonify([serialize_customer(row) for row in rows])


@app.post("/api/customers")
def create_customer():
  payload = request.get_json(force=True) or {}
  try:
    with db() as conn:
      customer_id = upsert_customer(conn, payload)
      row = fetch_customer(conn, customer_id)
    return jsonify({"ok": True, "customer": serialize_customer(row)})
  except Exception as exc:
    return json_error(str(exc))


@app.patch("/api/customers/<int:customer_id>")
def update_customer(customer_id: int):
  payload = request.get_json(force=True) or {}
  try:
    with db() as conn:
      row = fetch_customer(conn, customer_id)
      if not row:
        return json_error("Customer not found", 404)
      data = serialize_customer(row)
      data.update(payload)
      data["id"] = customer_id
      updated_id = upsert_customer(conn, data)
      updated = fetch_customer(conn, updated_id)
    return jsonify({"ok": True, "customer": serialize_customer(updated)})
  except Exception as exc:
    return json_error(str(exc))


@app.delete("/api/customers/<int:customer_id>")
def delete_customer(customer_id: int):
  with db() as conn:
    doc = conn.execute("SELECT id FROM documents WHERE customer_id=? LIMIT 1", (customer_id,)).fetchone()
    if doc:
      return json_error("Customer is used by documents and cannot be deleted", 400)
    conn.execute("DELETE FROM customers WHERE id=?", (customer_id,))
  return jsonify({"ok": True})


@app.get("/api/products")
def list_products():
  active_only = request.args.get("active_only") == "1"
  q = "SELECT * FROM products"
  if active_only:
    q += " WHERE is_active=1"
  q += " ORDER BY lower(name)"
  with db() as conn:
    rows = conn.execute(q).fetchall()
  return jsonify([serialize_product(row) for row in rows])


@app.post("/api/products")
def create_product():
  payload = request.get_json(force=True) or {}
  try:
    with db() as conn:
      product_id = upsert_product(conn, payload)
      row = fetch_product(conn, product_id)
    return jsonify({"ok": True, "product": serialize_product(row)})
  except Exception as exc:
    return json_error(str(exc))


@app.patch("/api/products/<int:product_id>")
def update_product(product_id: int):
  payload = request.get_json(force=True) or {}
  try:
    with db() as conn:
      row = fetch_product(conn, product_id)
      if not row:
        return json_error("Product not found", 404)
      data = serialize_product(row)
      data.update(payload)
      updated_id = upsert_product(conn, data)
      updated = fetch_product(conn, updated_id)
    return jsonify({"ok": True, "product": serialize_product(updated)})
  except Exception as exc:
    return json_error(str(exc))


@app.delete("/api/products/<int:product_id>")
def delete_product(product_id: int):
  with db() as conn:
    line = conn.execute("SELECT id FROM document_lines WHERE product_id=? LIMIT 1", (product_id,)).fetchone()
    if line:
      return json_error("Product is used by documents and cannot be deleted", 400)
    conn.execute("DELETE FROM products WHERE id=?", (product_id,))
  return jsonify({"ok": True})


@app.get("/api/documents")
def list_documents():
  doc_type = request.args.get("type")
  q = """SELECT d.id, d.type, d.number, d.issue_date, d.due_date, d.status,
                d.subtotal, d.tax_rate, d.tax_amount, d.total, d.imported,
                d.source_system, d.source_id, d.payment_url, d.last_sent_at, d.last_sent_to,
                c.name AS customer_name
         FROM documents d
         JOIN customers c ON c.id = d.customer_id
         WHERE 1=1"""
  params = []
  if doc_type in {"estimate", "invoice"}:
    q += " AND d.type=?"
    params.append(doc_type)
  q += " ORDER BY d.issue_date DESC, d.id DESC"
  with db() as conn:
    rows = conn.execute(q, params).fetchall()
  items = []
  for row in rows:
    items.append({
      "id": int(row["id"]),
      "type": row["type"],
      "number": row["number"],
      "issue_date": row["issue_date"],
      "due_date": row["due_date"],
      "status": row["status"],
      "subtotal": float(row["subtotal"] or 0),
      "tax_rate": float(row["tax_rate"] or 0),
      "tax_amount": float(row["tax_amount"] or 0),
      "total": float(row["total"] or 0),
      "imported": bool(row["imported"]),
      "source_system": row["source_system"],
      "source_id": row["source_id"],
      "customer_name": row["customer_name"],
      "payment_url": row["payment_url"] if "payment_url" in row.keys() else None,
      "last_sent_at": row["last_sent_at"] if "last_sent_at" in row.keys() else None,
      "last_sent_to": row["last_sent_to"] if "last_sent_to" in row.keys() else None,
    })
  return jsonify(items)


@app.get("/api/documents/<int:document_id>")
def get_document(document_id: int):
  with db() as conn:
    document = fetch_document(conn, document_id)
    if document:
      settings = business_settings_dict(conn)
      ensure_document_payment_url(conn, document, settings, persist=False)
  if not document:
    return json_error("Document not found", 404)
  return jsonify(document)


@app.get("/api/documents/<int:document_id>/email_draft")
def get_document_email_draft(document_id: int):
  with db() as conn:
    document = fetch_document(conn, document_id)
    if not document:
      return json_error("Document not found", 404)
    if document["type"] != "invoice":
      return json_error("Only invoices can be sent by email", 400)
    settings = business_settings_dict(conn)
    ensure_document_payment_url(conn, document, settings, persist=False)
    draft = build_invoice_email_draft(document, settings)
  return jsonify({"ok": True, "draft": draft, "document": document})


@app.post("/api/documents/<int:document_id>/send_email")
def send_document_email(document_id: int):
  payload = request.get_json(force=True) or {}
  try:
    with db() as conn:
      document = fetch_document(conn, document_id)
      if not document:
        return json_error("Document not found", 404)
      if document["type"] != "invoice":
        return json_error("Only invoices can be sent by email", 400)
      settings = business_settings_dict(conn)
      ensure_document_payment_url(conn, document, settings, persist=False)
      draft = build_invoice_email_draft(document, settings)
      to_email = clean_text(payload.get("to") or draft["to"])
      subject = clean_text(payload.get("subject") or draft["subject"])
      html_body = payload.get("html") or draft["html"]
      text_body = payload.get("text") or draft["text"]
      pdf_bytes = render_document_pdf_bytes(document, settings)
      send_invoice_email_message(
        settings=settings,
        to_email=to_email,
        subject=subject,
        html_body=html_body,
        text_body=text_body,
        pdf_name=f"{document['number']}.pdf",
        pdf_bytes=pdf_bytes,
        inline_images=draft.get("inline_images"),
      )
      try:
        conn.execute(
          "UPDATE documents SET payment_url=?, last_sent_at=?, last_sent_to=?, last_email_error=?, updated_at=? WHERE id=?",
          (document["payment_url"], now_iso(), to_email, None, now_iso(), document_id)
        )
      except Exception:
        pass
    return jsonify({"ok": True, "message": f"Invoice emailed to {to_email}"})
  except Exception as exc:
    try:
      with db() as conn:
        conn.execute(
          "UPDATE documents SET last_email_error=?, updated_at=? WHERE id=?",
          (str(exc), now_iso(), document_id)
        )
    except Exception:
      pass
    return json_error(str(exc))


@app.post("/api/documents")
def create_document():
  payload = request.get_json(force=True) or {}
  try:
    with db() as conn:
      document_id = save_document(conn, payload)
      document = fetch_document(conn, document_id)
    return jsonify({"ok": True, "document": document})
  except Exception as exc:
    return json_error(str(exc))


@app.patch("/api/documents/<int:document_id>")
def update_document(document_id: int):
  payload = request.get_json(force=True) or {}
  try:
    with db() as conn:
      if not fetch_document(conn, document_id):
        return json_error("Document not found", 404)
      save_document(conn, payload, document_id=document_id)
      document = fetch_document(conn, document_id)
    return jsonify({"ok": True, "document": document})
  except Exception as exc:
    return json_error(str(exc))


@app.delete("/api/documents/<int:document_id>")
def delete_document(document_id: int):
  with db() as conn:
    conn.execute("DELETE FROM document_lines WHERE document_id=?", (document_id,))
    conn.execute("DELETE FROM documents WHERE id=?", (document_id,))
  return jsonify({"ok": True})


@app.post("/api/documents/<int:document_id>/convert_to_invoice")
def convert_document_to_invoice(document_id: int):
  try:
    with db() as conn:
      document = fetch_document(conn, document_id)
      if not document:
        return json_error("Estimate not found", 404)
      if document["type"] != "estimate":
        return json_error("Only estimates can be converted", 400)
      invoice_payload = {
        "type": "invoice",
        "customer_id": document["customer_id"],
        "issue_date": date.today().isoformat(),
        "due_date": document.get("due_date"),
        "status": "open",
        "tax_rate": document["tax_rate"],
        "notes": document["notes"],
        "terms": document["terms"],
        "converted_from_document_id": document["id"],
        "lines": [
          {
            "product_id": line.get("product_id"),
            "description": line["description"],
            "quantity": line["quantity"],
            "unit_price": line["unit_price"],
            "taxable": line["taxable"],
          }
          for line in document["lines"]
        ],
      }
      invoice_id = save_document(conn, invoice_payload)
      invoice = fetch_document(conn, invoice_id)
    return jsonify({"ok": True, "document": invoice})
  except Exception as exc:
    return json_error(str(exc))


@app.get("/api/documents/<int:document_id>/print")
def print_document(document_id: int):
  with db() as conn:
    document = fetch_document(conn, document_id)
    if not document:
      return json_error("Document not found", 404)
    settings = business_settings_dict(conn)
  return render_template(
    "document_print.html",
    title=document["number"],
    document=document,
    business=settings,
    logo_src="/api/business_logo",
  )


@app.get("/api/documents/<int:document_id>/pdf")
def pdf_document(document_id: int):
  try:
    with db() as conn:
      document = fetch_document(conn, document_id)
      if not document:
        return json_error("Document not found", 404)
      settings = business_settings_dict(conn)
    payload = render_document_pdf_bytes(document, settings)
    return send_file(
      io.BytesIO(payload),
      mimetype="application/pdf",
      as_attachment=True,
      download_name=f"{document['number']}.pdf"
    )
  except Exception as exc:
    return json_error(str(exc))


@app.get("/pay/<int:document_id>")
def pay_document(document_id: int):
  with db() as conn:
    document = fetch_document(conn, document_id)
    if not document:
      return json_error("Document not found", 404)
    settings = business_settings_dict(conn)
    ensure_document_payment_url(conn, document, settings, persist=False)
  payment_base = clean_text(settings.get("invoice_payment_url_base"))
  external_payment_url = None
  if payment_base:
    external_payment_url = build_payment_url(document, settings)
  return render_template(
    "invoice_payment.html",
    document=document,
    business=settings,
    external_payment_url=external_payment_url,
    sbooks_logo_src=sbooks_brand_data_uri(),
  )


@app.post("/api/import/quickbooks/preview")
def preview_quickbooks_import():
  try:
    customers_file = request.files.get("customers_file")
    products_file = request.files.get("products_file")
    invoices_file = request.files.get("invoices_file")
    backup_file = request.files.get("backup_file")

    if backup_file and backup_file.filename:
      preview = preview_quickbooks_bundle(backup_file)
    else:
      customers, c_warn, c_err = normalize_customer_rows(read_table_file(customers_file))
      products, p_warn, p_err = normalize_product_rows(read_table_file(products_file))
      documents, d_warn, d_err = normalize_invoice_rows(read_table_file(invoices_file))
      preview = {
        "customers": customers,
        "products": products,
        "documents": documents,
        "warnings": [*c_warn, *p_warn, *d_warn],
        "errors": [*c_err, *p_err, *d_err],
      }
    preview["summary"] = {
      "customers": len(preview["customers"]),
      "products": len(preview["products"]),
      "documents": len(preview["documents"]),
    }
    return jsonify(preview)
  except Exception as exc:
    return json_error(str(exc))


@app.post("/api/import/quickbooks/commit")
def commit_quickbooks_import():
  payload = request.get_json(force=True) or {}
  preview = payload.get("preview") or {}
  customers = preview.get("customers") or []
  products = preview.get("products") or []
  documents = preview.get("documents") or []
  try:
    with db() as conn:
      customer_map_by_name = {}
      product_map_by_name = {}
      for item in customers:
        customer_id = upsert_customer(conn, item)
        customer_map_by_name[clean_text(item.get("name")).lower()] = customer_id
      for item in products:
        product_id = upsert_product(conn, item)
        product_map_by_name[clean_text(item.get("name")).lower()] = product_id

      imported_ids = []
      for item in documents:
        customer_name = clean_text(item.get("customer_name")) or "Imported Customer"
        customer_id = customer_map_by_name.get(customer_name.lower())
        if not customer_id:
          customer_id = upsert_customer(conn, {
            "name": customer_name,
            "external_source": "quickbooks",
            "external_id": None,
          })
          customer_map_by_name[customer_name.lower()] = customer_id

        lines = []
        for line in item.get("lines", []):
          product_id = None
          product_name = clean_text(line.get("product_name"))
          if product_name:
            product_id = product_map_by_name.get(product_name.lower())
            if not product_id:
              product_id = upsert_product(conn, {
                "name": product_name,
                "description": line.get("description") or product_name,
                "default_unit_price": line.get("unit_price", 0),
                "taxable": line.get("taxable", True),
                "external_source": "quickbooks",
                "external_id": None,
              })
              product_map_by_name[product_name.lower()] = product_id
          lines.append({
            "product_id": product_id,
            "description": line.get("description") or product_name or "Imported line item",
            "quantity": line.get("quantity", 1),
            "unit_price": line.get("unit_price", 0),
            "taxable": line.get("taxable", True),
          })

        source_system = clean_text(item.get("source_system")) or "quickbooks"
        source_id = clean_text(item.get("source_id")) or clean_text(item.get("number"))
        existing = None
        if source_system and source_id:
          existing = conn.execute(
            "SELECT id FROM documents WHERE source_system=? AND source_id=?",
            (source_system, source_id)
          ).fetchone()
        document_payload = {
          "type": item.get("type") or "invoice",
          "number": item.get("number"),
          "customer_id": customer_id,
          "issue_date": item.get("issue_date"),
          "due_date": item.get("due_date"),
          "status": item.get("status") or "open",
          "tax_rate": item.get("tax_rate", 0),
          "notes": item.get("notes"),
          "terms": item.get("terms"),
          "imported": True,
          "source_system": source_system,
          "source_id": source_id,
          "lines": lines,
        }
        document_id = save_document(conn, document_payload, int(existing["id"]) if existing else None)
        imported_ids.append(document_id)
    return jsonify({"ok": True, "summary": {"customers": len(customers), "products": len(products), "documents": len(imported_ids)}, "document_ids": imported_ids})
  except Exception as exc:
    return json_error(str(exc))


@app.get("/api/archives")
def list_archives():
  """List archived pay periods (by start_date)."""
  with db() as conn:
    try:
      rows = conn.execute(
        "SELECT start_date, archived_at FROM pay_period_archives ORDER BY start_date DESC"
      ).fetchall()
      items = [dict(r) for r in rows]
    except Exception:
      items = []
  return jsonify({"archives": items})


@app.post("/api/pay_periods/<start_date>/archive")
def archive_pay_period(start_date: str):
  """Archive a pay period by its start_date (YYYY-MM-DD)."""
  # Validate
  try:
    _ = parse_date(start_date)
  except Exception:
    return jsonify({"ok": False, "error": "Invalid start_date"}), 400
  with db() as conn:
    try:
      conn.execute(
        "INSERT OR IGNORE INTO pay_period_archives(start_date, archived_at) VALUES (?,?)",
        (start_date, now_iso()),
      )
    except Exception as e:
      return jsonify({"ok": False, "error": str(e)}), 500
  return jsonify({"ok": True})


@app.delete("/api/pay_periods/<start_date>/archive")
def unarchive_pay_period(start_date: str):
  """Undo archive for a pay period by its start_date."""
  try:
    _ = parse_date(start_date)
  except Exception:
    return jsonify({"ok": False, "error": "Invalid start_date"}), 400
  with db() as conn:
    try:
      conn.execute("DELETE FROM pay_period_archives WHERE start_date=?", (start_date,))
    except Exception as e:
      return jsonify({"ok": False, "error": str(e)}), 500
  return jsonify({"ok": True})

@app.post("/api/accounts/anchor")
def set_anchor():
  payload = request.get_json(force=True)
  account_id = int(payload["account_id"])
  anchor_date = payload["anchor_date"]
  anchor_balance = float(payload["anchor_balance"])
  with db() as conn:
    conn.execute(
      "INSERT OR REPLACE INTO account_anchors(account_id, anchor_date, anchor_balance) VALUES (?,?,?)",
      (account_id, anchor_date, anchor_balance)
    )
  return jsonify({"ok": True})

@app.get("/api/recurring_rules")
def list_rules():
  with db() as conn:
    rules = [dict(r) for r in conn.execute("SELECT * FROM recurring_rules ORDER BY id DESC").fetchall()]
  return jsonify(rules)

@app.post("/api/recurring_rules")
def create_rule():
  p = request.get_json(force=True)
  url = (p.get("url") or "").strip() or None
  with db() as conn:
    cur = conn.execute(
      """INSERT INTO recurring_rules
         (account_id, description, url, amount, cadence, day_of_month, by_day_of_month, start_date, end_date, due_day, funding_bucket_override, is_active)
         VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
      (
        int(p["account_id"]),
        p["description"],
        url,
        float(p["amount"]),
        p["cadence"],
        p.get("day_of_month"),
        int(bool(p.get("by_day_of_month", False))),
        p["start_date"],
        p.get("end_date"),
        p.get("due_day"),
        p.get("funding_bucket_override"),
        1
      )
    )
    rule_id = cur.lastrowid
  return jsonify({"ok": True, "id": rule_id})


@app.patch("/api/recurring_rules/<int:rule_id>")
def update_rule(rule_id: int):
  p = request.get_json(force=True)
  allowed = {
    "account_id",
    "description",
    "url",
    "amount",
    "cadence",
    "day_of_month",
    "by_day_of_month",
    "start_date",
    "end_date",
    "due_day",
    "funding_bucket_override",
    "is_active",
  }
  fields = []
  params = []
  for key, val in p.items():
    if key not in allowed:
      continue
    if key == "url":
      val = (val or "").strip() or None
    if key == "by_day_of_month":
      val = int(bool(val))
    if key == "is_active":
      val = int(bool(val))
    fields.append(f"{key}=?")
    params.append(val)
  if not fields:
    return jsonify({"ok": True})
  params.append(rule_id)
  with db() as conn:
    row = conn.execute("SELECT id FROM recurring_rules WHERE id=?", (rule_id,)).fetchone()
    if not row:
      return jsonify({"ok": False, "error": "Rule not found"}), 404
    conn.execute(f"UPDATE recurring_rules SET {', '.join(fields)} WHERE id=?", params)
  return jsonify({"ok": True})


@app.delete("/api/recurring_rules/<int:rule_id>")
def delete_rule(rule_id: int):
  with db() as conn:
    conn.execute("DELETE FROM recurring_rules WHERE id=?", (rule_id,))
  return jsonify({"ok": True})

@app.post("/api/recurring_rules/<int:rule_id>/generate")
def generate_from_rule(rule_id: int):
  p = request.get_json(force=True)
  to_date = parse_date(p["to_date"])
  with db() as conn:
    rule = conn.execute("SELECT * FROM recurring_rules WHERE id=?", (rule_id,)).fetchone()
    if not rule:
      return jsonify({"ok": False, "error": "Rule not found"}), 404

    start = parse_date(rule["start_date"])
    end = parse_date(rule["end_date"]) if rule["end_date"] else to_date
    end = min(end, to_date)

    paychecks = [dict(r) for r in conn.execute("SELECT id,date FROM paychecks ORDER BY date").fetchall()]

    created = 0
    if rule["cadence"] == "monthly":
      dom = int(rule["day_of_month"] or 1)
      occ = date(start.year, start.month, 1)
      while occ <= end:
        next_month = (occ.replace(day=28) + timedelta(days=4)).replace(day=1)
        last_day = (next_month - timedelta(days=1)).day
        day = min(dom, last_day)
        eff = date(occ.year, occ.month, day)
        if eff < start:
          occ = next_month
          continue
        if eff > end:
          break

        due_day = int(rule["due_day"] or day)
        bucket = rule["funding_bucket_override"] or bucket_from_due_day(due_day)
        fp_id = assign_funding_paycheck_id(eff.year, eff.month, bucket, paychecks)
        due_label = f"by the {due_day}th" if rule["by_day_of_month"] else f"{due_day}th"

        conn.execute(
          """INSERT INTO transactions
             (account_id, recurring_rule_id, effective_date, amount, description, url, due_day, due_label, funding_bucket, funding_paycheck_id, status, sort_key, created_at, updated_at)
             VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
          (
            rule["account_id"], rule["id"], iso(eff), float(rule["amount"]), rule["description"], (rule["url"] if ("url" in rule.keys()) else None),
            due_day, due_label, bucket, fp_id, "planned", 0, now_iso(), now_iso()
          )
        )
        created += 1
        occ = next_month

    elif rule["cadence"] == "biweekly":
      d = start
      while d <= end:
        due_day = int(rule["due_day"] or d.day)
        bucket = rule["funding_bucket_override"] or bucket_from_due_day(due_day)
        fp_id = assign_funding_paycheck_id(d.year, d.month, bucket, paychecks)

        conn.execute(
          """INSERT INTO transactions
             (account_id, recurring_rule_id, effective_date, amount, description, url, due_day, due_label, funding_bucket, funding_paycheck_id, status, sort_key, created_at, updated_at)
             VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
          (
            rule["account_id"], rule["id"], iso(d), float(rule["amount"]), rule["description"], (rule["url"] if ("url" in rule.keys()) else None),
            due_day, None, bucket, fp_id, "planned", 0, now_iso(), now_iso()
          )
        )
        created += 1
        d += timedelta(days=14)
    else:
      return jsonify({"ok": False, "error": "Cadence not implemented in MVP"}), 400

  return jsonify({"ok": True, "created": created})

@app.get("/api/transactions")
def list_transactions():
  account_id = request.args.get("account_id")
  date_from = request.args.get("from")
  date_to = request.args.get("to")
  q = "SELECT * FROM transactions WHERE 1=1"
  params = []
  if account_id:
    q += " AND account_id=?"
    params.append(int(account_id))
  if date_from:
    q += " AND effective_date>=?"
    params.append(date_from)
  if date_to:
    q += " AND effective_date<=?"
    params.append(date_to)
  q += " ORDER BY effective_date, sort_key, id"
  with db() as conn:
    txs = [dict(r) for r in conn.execute(q, params).fetchall()]
  return jsonify(txs)

@app.post("/api/transactions")
def create_transaction():
  p = request.get_json(force=True)
  eff = parse_date(p["effective_date"])
  due_day_raw = p.get("due_day")
  due_day = int(due_day_raw) if due_day_raw not in (None, "", "null") else None
  bucket = p.get("funding_bucket") or bucket_from_due_day(due_day)
  url = (p.get("url") or "").strip() or None

  with db() as conn:
    paychecks = [dict(r) for r in conn.execute("SELECT id,date FROM paychecks ORDER BY date").fetchall()]
    fp_id = p.get("funding_paycheck_id")
    if fp_id in (None, "", "null"):
      fp_id = assign_funding_paycheck_id(eff.year, eff.month, bucket, paychecks)

    cur = conn.execute(
      """INSERT INTO transactions
         (account_id, recurring_rule_id, effective_date, amount, description, url, due_day, due_label, funding_bucket, funding_paycheck_id, status, sort_key, created_at, updated_at)
         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
      (
        int(p["account_id"]),
        p.get("recurring_rule_id"),
        iso(eff),
        float(p["amount"]),
        p["description"],
        url,
        due_day,
        p.get("due_label"),
        bucket,
        int(fp_id) if fp_id else None,
        p.get("status", "planned"),
        int(p.get("sort_key", 0)),
        now_iso(),
        now_iso()
      )
    )
    tx_id = cur.lastrowid
  return jsonify({"ok": True, "id": tx_id})

@app.patch("/api/transactions/<int:tx_id>")
def update_transaction(tx_id: int):
  p = request.get_json(force=True)
  allowed = {"account_id","effective_date","amount","description","url","due_day","due_label","funding_bucket","funding_paycheck_id","status","sort_key"}
  fields = []
  params = []
  for key, val in p.items():
    if key in allowed:
      fields.append(f"{key}=?")
      params.append(val)
  fields.append("updated_at=?")
  params.append(now_iso())
  params.append(tx_id)
  with db() as conn:
    conn.execute(f"UPDATE transactions SET {', '.join(fields)} WHERE id=?", params)
  return jsonify({"ok": True})

@app.delete("/api/transactions/<int:tx_id>")
def delete_transaction(tx_id: int):
  with db() as conn:
    conn.execute("DELETE FROM transactions WHERE id=?", (tx_id,))
  return jsonify({"ok": True})

@app.post("/api/transactions/recompute_funding")
def recompute_funding():
  p = request.get_json(force=True)
  date_from = p.get("from", "1900-01-01")
  date_to = p.get("to", "2999-12-31")
  with db() as conn:
    paychecks = [dict(r) for r in conn.execute("SELECT id,date FROM paychecks ORDER BY date").fetchall()]
    txs = conn.execute(
      "SELECT id,effective_date,due_day FROM transactions WHERE effective_date>=? AND effective_date<=? ORDER BY id",
      (date_from, date_to)
    ).fetchall()
    updated = 0
    for r in txs:
      eff = parse_date(r["effective_date"])
      bucket = bucket_from_due_day(r["due_day"])
      fp_id = assign_funding_paycheck_id(eff.year, eff.month, bucket, paychecks)
      conn.execute("UPDATE transactions SET funding_bucket=?, funding_paycheck_id=?, updated_at=? WHERE id=?",
                   (bucket, fp_id, now_iso(), r["id"]))
      updated += 1
  return jsonify({"ok": True, "updated": updated})



@app.patch("/api/cc_snapshots/<int:snapshot_id>")
def update_cc_snapshot(snapshot_id: int):
  p = request.get_json(force=True)
  name = (p.get("name") or "").strip()
  snapshot_date = (p.get("snapshot_date") or "").strip()
  side = (p.get("side") or "personal").strip()
  url = (p.get("url") or "").strip() or None
  # Card due day is card-level metadata and only applies to card-definition rows.
  due_day = (p.get("due_day") or "").strip() or None
  pay_status = (p.get("pay_status") or "").strip() or None
  try:
    balance = float(p.get("balance") or 0)
  except Exception:
    balance = 0.0

  if not name:
    return jsonify({"ok": False, "error": "Name is required"}), 400
  if not snapshot_date:
    return jsonify({"ok": False, "error": "Snapshot date is required"}), 400

  with db() as conn:
    row = conn.execute("SELECT id FROM cc_snapshots WHERE id=?", (snapshot_id,)).fetchone()
    if not row:
      return jsonify({"ok": False, "error": "Not found"}), 404
    cols = [r[1] for r in conn.execute("PRAGMA table_info(cc_snapshots)").fetchall()]
    has_pay_status = "pay_status" in cols

    # Only allow editing due_day on card-definition rows.
    if snapshot_date == CC_CARDDEF_DATE:
      conn.execute(
        "UPDATE cc_snapshots SET name=?, snapshot_date=?, balance=?, side=?, url=?, due_day=? WHERE id=?",
        (name, snapshot_date, balance, side, url, due_day, snapshot_id)
      )
    else:
      if has_pay_status:
        conn.execute(
          "UPDATE cc_snapshots SET name=?, snapshot_date=?, balance=?, side=?, url=?, pay_status=? WHERE id=?",
          (name, snapshot_date, balance, side, url, pay_status, snapshot_id)
        )
      else:
        conn.execute(
          "UPDATE cc_snapshots SET name=?, snapshot_date=?, balance=?, side=?, url=? WHERE id=?",
          (name, snapshot_date, balance, side, url, snapshot_id)
        )
  return jsonify({"ok": True})

@app.delete("/api/cc_snapshots/<int:snapshot_id>")
def delete_cc_snapshot(snapshot_id: int):
  with db() as conn:
    conn.execute("DELETE FROM cc_snapshots WHERE id=?", (snapshot_id,))
  return jsonify({"ok": True})

# ---------------- Credit Card Snapshots ----------------

@app.get("/api/cc_snapshots")
def list_cc_snapshots():
  as_of = request.args.get("as_of")
  side = request.args.get("side") or "personal"
  with db() as conn:
    if as_of:
      cards, total = cc_cards_for_paycheck(conn, as_of, side=side)
      return jsonify({"as_of": as_of, "cards": cards, "total": total})
    rows = conn.execute("SELECT * FROM cc_snapshots WHERE side=? ORDER BY snapshot_date DESC, name", (side,)).fetchall()
    return jsonify([dict(r) for r in rows])

@app.post("/api/cc_snapshots/save")
def save_cc_snapshots():
  p = request.get_json(force=True)
  snap_date = p.get("snapshot_date") or date.today().isoformat()
  side = p.get("side") or "personal"
  cards = p.get("cards", [])
  with db() as conn:
    for c in cards:
      name = (c.get("name") or "").strip()
      if not name:
        continue
      url = (c.get("url") or "").strip() or None
      bal = float(c.get("balance") or 0)
      pay_status = (c.get("pay_status") or "").strip() or None
      # Keep one row per (name, snapshot_date, side)
      conn.execute(
        "DELETE FROM cc_snapshots WHERE name=? AND snapshot_date=? AND side=?",
        (name, snap_date, side)
      )
      conn.execute(
        "INSERT INTO cc_snapshots(name, snapshot_date, balance, side, url, pay_status) VALUES (?,?,?,?,?,?)",
        (name, snap_date, bal, side, url, pay_status)
      )
  return jsonify({"ok": True})


# ---------------- Card Definitions (name/url/due_day) ----------------

@app.get("/api/cc_cards")
def list_cc_cards():
  side = request.args.get("side") or "personal"
  with db() as conn:
    defs = cc_card_defs(conn, side)
    # If no explicit card-def rows exist yet, infer from existing snapshots.
    if not defs:
      names = cc_card_names(conn, side)
      today_iso = date.today().isoformat()
      defs = []
      for nm in names:
        defs.append({
          "id": None,
          "name": nm,
          "url": cc_latest_url(conn, nm, today_iso, side),
          "due_day": cc_latest_due_day(conn, nm, today_iso, side)
        })
    return jsonify({"side": side, "cards": defs})


def _normalize_due_day(v: str | None):
  if not v:
    return None
  s = str(v).strip()
  if not s:
    return None
  # Accept '5' -> '05'
  if s.isdigit() and len(s) in (1, 2):
    n = int(s)
    if 1 <= n <= 31:
      return f"{n:02d}"
  return None


@app.post("/api/cc_cards/save")
def save_cc_cards():
  p = request.get_json(force=True)
  side = p.get("side") or "personal"
  cards = p.get("cards", [])
  with db() as conn:
    # Rewrite card-definition rows for this side (non-destructive to pay-period snapshots).
    conn.execute("DELETE FROM cc_snapshots WHERE side=? AND snapshot_date=?", (side, CC_CARDDEF_DATE))
    for c in cards:
      name = (c.get("name") or "").strip()
      if not name:
        continue
      url = (c.get("url") or "").strip() or None
      due_day = _normalize_due_day(c.get("due_day"))
      conn.execute(
        "INSERT INTO cc_snapshots(name, snapshot_date, balance, side, url, due_day) VALUES (?,?,?,?,?,?)",
        (name, CC_CARDDEF_DATE, 0.0, side, url, due_day)
      )
  return jsonify({"ok": True})

@app.post("/api/cc/create_payment")
def create_cc_payment():
  p = request.get_json(force=True)
  period_start = parse_date(p["period_start"])
  offset_days = int(p.get("offset_days", 6))
  side = p.get("side") or "personal"
  account_id = int(p.get("account_id", 1 if side=="personal" else 2))
  desc = p.get("description", "cc" if side=="personal" else "cc-biz")
  eff = period_start + timedelta(days=offset_days)

  with db() as conn:
    cards, total = cc_cards_for_paycheck(conn, iso(period_start), side=side)
    if total == 0:
      return jsonify({"ok": False, "error": "CC total is $0.00 (no snapshots as of that date)"}), 400

    existing = conn.execute(
      "SELECT id FROM transactions WHERE account_id=? AND effective_date=? AND description=?",
      (account_id, iso(eff), desc)
    ).fetchone()
    if existing:
      return jsonify({"ok": True, "id": existing["id"], "skipped": True, "amount": -total})

    cur = conn.execute(
      """INSERT INTO transactions
         (account_id, recurring_rule_id, effective_date, amount, description, url, due_day, due_label, funding_bucket, funding_paycheck_id, status, sort_key, created_at, updated_at)
         VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
      (
        account_id, None, iso(eff), float(-total), desc, None,
        None, None, "ONE_OFF", None, "planned", 50, now_iso(), now_iso()
      )
    )
    return jsonify({"ok": True, "id": cur.lastrowid, "amount": -total, "cards": cards, "total": total})

@app.get("/api/projection")
def projection():
  q_from = request.args.get("from")
  q_to = request.args.get("to")

  with db() as conn:
    settings = {r["key"]: r["value"] for r in conn.execute("SELECT key,value FROM settings").fetchall()}
    anchor_date = parse_date(settings.get("anchor_date", date.today().isoformat()))
    horizon_days = int(settings.get("horizon_days", "365"))
    to_default = anchor_date + timedelta(days=horizon_days)

    date_from = parse_date(q_from) if q_from else anchor_date
    date_to = parse_date(q_to) if q_to else to_default

    accounts = [dict(r) for r in conn.execute("SELECT * FROM accounts ORDER BY id").fetchall()]
    paychecks = [dict(r) for r in conn.execute("SELECT id,account_id,date,amount FROM paychecks ORDER BY date").fetchall()]

    primary_account_id = int(settings.get("paycheck_account_id", "1"))
    primary = [p for p in paychecks if p["account_id"] == primary_account_id]
    primary.sort(key=lambda p: p["date"])

    archived_starts = _archived_period_starts(conn)

    periods = []
    for i, p in enumerate(primary):
      start = parse_date(p["date"])
      end = parse_date(primary[i+1]["date"]) - timedelta(days=1) if i + 1 < len(primary) else (start + timedelta(days=13))
      if end < date_from or start > date_to:
        continue
      if iso(start) in archived_starts:
        continue
      periods.append({"start_date": iso(start), "end_date": iso(end), "paycheck": p})

    count_by_month = {}
    for p in primary:
      d = parse_date(p["date"])
      key = f"{d.year:04d}-{d.month:02d}"
      count_by_month[key] = count_by_month.get(key, 0) + 1

    tx_rows = conn.execute(
      """SELECT * FROM transactions
         WHERE effective_date>=? AND effective_date<=?
         ORDER BY effective_date, sort_key, id""",
      (iso(date_from - timedelta(days=1)), iso(date_to + timedelta(days=1)))
    ).fetchall()
    txs = [dict(r) for r in tx_rows]

    anchors = conn.execute("SELECT * FROM account_anchors WHERE anchor_date=?", (settings.get("anchor_date", iso(anchor_date)),)).fetchall()
    anchor_by_account = {r["account_id"]: float(r["anchor_balance"]) for r in anchors}

    current_start = {a["id"]: anchor_by_account.get(a["id"], 0.0) for a in accounts}

    out_periods = []
    for period in periods:
      ps = parse_date(period["start_date"])
      pe = parse_date(period["end_date"])
      month_key = f"{ps.year:04d}-{ps.month:02d}"
      three_paycheck_month = (count_by_month.get(month_key, 0) == 3)

      period_out = {
        "start_date": period["start_date"],
        "end_date": period["end_date"],
        "three_paycheck_month": three_paycheck_month,
        "paycheck": period["paycheck"],
        "accounts": [],
        "cc": {"cards": [], "total": 0.0},
        "cc_biz": {"cards": [], "total": 0.0}
      }

      cc_cards, cc_total = cc_cards_for_paycheck(conn, iso(ps), side="personal")
      period_out["cc"] = {"cards": cc_cards, "total": cc_total}
      cc2_cards, cc2_total = cc_cards_for_paycheck(conn, iso(ps), side="business")
      period_out["cc_biz"] = {"cards": cc2_cards, "total": cc2_total}

      for a in accounts:
        acct_txs = [t for t in txs if t["account_id"] == a["id"] and ps <= parse_date(t["effective_date"]) <= pe]
        paycheck_tx = None
        if period["paycheck"]["account_id"] == a["id"]:
          paycheck_tx = {
            "id": f"paycheck-{period['paycheck']['id']}",
            "effective_date": period["start_date"],
            "amount": float(period["paycheck"]["amount"]),
            "description": "Paycheck",
            "due_label": None,
            "status": "planned",
            "sort_key": -999
          }

        tx_by_day = {}
        for t in acct_txs:
          tx_by_day.setdefault(t["effective_date"], []).append(t)
        if paycheck_tx:
          tx_by_day.setdefault(paycheck_tx["effective_date"], []).insert(0, paycheck_tx)

        bal = float(current_start[a["id"]])
        days = []
        for d in daterange(ps, pe):
          ds = iso(d)
          items = tx_by_day.get(ds, [])
          items_sorted = sorted(items, key=lambda x: (int(x.get("sort_key", 0)), str(x.get("id"))))
          start_bal = bal
          for it in items_sorted:
            bal += float(it["amount"])
          days.append({"date": ds, "items": items_sorted, "balance": round(bal, 2), "start_balance": round(start_bal, 2)})

        current_start[a["id"]] = bal
        start_balance = days[0]["start_balance"] if days else round(bal, 2)

        period_out["accounts"].append({
          "account_id": a["id"],
          "account_name": a["name"],
          "start_balance": round(start_balance, 2),
          "end_balance": round(bal, 2),
          "days": days
        })

      out_periods.append(period_out)

  return jsonify({"settings": settings, "accounts": accounts, "periods": out_periods})


@app.route("/api/export/google_sheet", methods=["POST"])
def export_google_sheet():
    r"""
    Export pay periods to Google Sheets using OAuth.
    Uses OAuth client file located in:
    C:\projects\cash16\oauth.json
    """

    import pickle
    from flask import jsonify
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    from google.auth.transport.requests import Request
    import os

    SHEET_ID = "1j7Gho1O55eCQl6Ca9wjksUkHHNOlBP89OUf6c-joQB8"
    CLIENT_SECRET = r"C:\projects\cash16\google_oauth.json"
    SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]

    creds = None

    if os.path.exists("token.pickle"):
        with open("token.pickle", "rb") as token:
            creds = pickle.load(token)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET, SCOPES)
            creds = flow.run_local_server(port=0)

        with open("token.pickle", "wb") as token:
            pickle.dump(creds, token)

    try:
        service = build("sheets", "v4", credentials=creds)

        with db() as conn:
            rows = conn.execute(
                "SELECT id, account_id, effective_date, amount, description FROM transactions ORDER BY effective_date"
            ).fetchall()

        values = [["ID","Account","Date","Amount","Description"]]
        for r in rows:
            values.append([
                r["id"],
                r["account_id"],
                r["effective_date"],
                float(r["amount"]),
                r["description"]
            ])

        body = {"values": values}

        service.spreadsheets().values().update(
            spreadsheetId=SHEET_ID,
            range="Transactions!A1",
            valueInputOption="RAW",
            body=body
        ).execute()

        return jsonify({"status":"ok","rows":len(values)})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


def main():
  init_db()
  app.run(host="127.0.0.1", port=5000, debug=False)

if __name__ == "__main__":
  main()
